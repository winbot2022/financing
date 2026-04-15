import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import io
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
from datetime import datetime
from openpyxl import load_workbook

st.set_page_config(page_title="資金繰り表 作成アプリ", layout="wide")

from pathlib import Path
from matplotlib import font_manager, rcParams

def setup_japanese_font():
    candidates = [
        Path(__file__).with_name("NotoSansJP-Regular.ttf"),
        Path.cwd() / "NotoSansJP-Regular.ttf",
    ]
    for font_path in candidates:
        if font_path.exists():
            try:
                font_manager.fontManager.addfont(str(font_path))
                font_name = font_manager.FontProperties(fname=str(font_path)).get_name()
                rcParams["font.family"] = font_name
                rcParams["axes.unicode_minus"] = False
                return
            except Exception:
                pass

def export_cashflow_template(input_rows, carry_in, start_month, template_path="資金繰り表完コピ版.xlsx"):
    wb = load_workbook(template_path)
    ws = wb["資金繰り表"]

    row_map = {
        "現金売上": 5,
        "売掛金回収": 6,
        "手形期日落": 7,
        "手形割引": 8,
        "前受金": 9,
        "その他収入": 10,
        "現金仕入": 12,
        "買掛金支払": 13,
        "手形決済": 14,
        "賃金及び給与": 15,
        "家賃": 16,
        "前渡金": 17,
        "諸経費": 18,
        "その他（設備等）": 19,
        "借入金返済": 22,
        "借入金": 23,
    }

    months_all = [f"{i}月" for i in range(1, 13)]
    start_idx = months_all.index(start_month)
    target_months = [months_all[(start_idx + i) % 12] for i in range(6)]

    # 月見出しを書き換え
    for i, m in enumerate(target_months):
        col = 4 + i * 2
        ws.cell(row=2, column=col, value=m)

    # 初月前月繰越
    ws["D4"] = carry_in

    if input_rows is not None and not input_rows.empty:
        grouped = input_rows.groupby(["月", "小分類"], as_index=False)["金額"].sum()

        for _, r in grouped.iterrows():
            month_name = r["月"]
            item = r["小分類"]
            amount = r["金額"]

            if month_name in target_months and item in row_map:
                month_idx = target_months.index(month_name)
                col = 4 + month_idx * 2
                row = row_map[item]
                ws.cell(row=row, column=col, value=float(amount))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def restore_cashflow_template(file):
    wb = load_workbook(file, data_only=False)
    ws = wb["資金繰り表"]

    row_map = {
        5: "現金売上",
        6: "売掛金回収",
        7: "手形期日落",
        8: "手形割引",
        9: "前受金",
        10: "その他収入",
        12: "現金仕入",
        13: "買掛金支払",
        14: "手形決済",
        15: "賃金及び給与",
        16: "家賃",
        17: "前渡金",
        18: "諸経費",
        19: "その他（設備等）",
        22: "借入金返済",
        23: "借入金",
    }

    restored_rows = []

    carry_in = ws["D4"].value
    if isinstance(carry_in, str) and carry_in.startswith("="):
        carry_in = 0
    carry_in = int(carry_in or 0)

    # 月見出しを読む
    target_months = []
    for i in range(6):
        col = 4 + i * 2
        target_months.append(str(ws.cell(row=2, column=col).value).strip())

    start_month = target_months[0]

    for month_idx, month_name in enumerate(target_months):
        col = 4 + month_idx * 2

        for row_num, item_name in row_map.items():
            val = ws.cell(row=row_num, column=col).value
            if val is None:
                continue
            if isinstance(val, str) and val.startswith("="):
                continue
            if float(val) == 0:
                continue

            if item_name in ["現金売上", "売掛金回収", "手形期日落", "手形割引", "前受金", "その他収入"]:
                kubun = "収入"
                major_map = {
                    "現金売上": "売上代金",
                    "売掛金回収": "売上代金",
                    "手形期日落": "売上代金",
                    "手形割引": "売上代金",
                    "前受金": "前受金",
                    "その他収入": "その他収入",
                }
            elif item_name in ["借入金返済", "借入金"]:
                kubun = "財務"
                major_map = {
                    "借入金返済": "借入金返済",
                    "借入金": "借入金",
                }
            else:
                kubun = "支出"
                major_map = {
                    "現金仕入": "仕入代金",
                    "買掛金支払": "仕入代金",
                    "手形決済": "仕入代金",
                    "賃金及び給与": "賃金及び給与",
                    "家賃": "家賃",
                    "前渡金": "前渡金",
                    "諸経費": "諸経費",
                    "その他（設備等）": "その他（設備等）",
                }

            restored_rows.append({
                "月": month_name,
                "区分": kubun,
                "大分類": major_map[item_name],
                "小分類": item_name,
                "金額": int(val),
            })

    return start_month, carry_in, pd.DataFrame(restored_rows)

setup_japanese_font()

st.title("資金繰り表 作成アプリ")

MONTHS = [f"{i}月" for i in range(1, 13)]

CATEGORY_STRUCTURE = {
    "収入": {
        "売上代金": ["現金売上", "売掛金回収", "手形期日落", "手形割引"],
        "前受金": ["前受金"],
        "その他収入": ["その他収入"],
    },
    "支出": {
        "仕入代金": ["現金仕入", "買掛金支払", "手形決済"],
        "賃金及び給与": ["賃金及び給与"],
        "家賃": ["家賃"],
        "前渡金": ["前渡金"],
        "諸経費": ["諸経費"],
        "その他（設備等）": ["その他（設備等）"],
    },
    "財務": {
        "借入金": ["借入金"],
        "借入金返済": ["借入金返済"],
    },
}

income_rows = [
    "現金売上", "売掛金回収", "手形期日落", "手形割引",
    "前受金", "その他収入"
]

expense_rows = [
    "現金仕入", "買掛金支払", "手形決済",
    "賃金及び給与", "家賃", "前渡金", "諸経費", "その他（設備等）"
]

finance_rows = [
    "借入金", "借入金返済"
]

DISPLAY_ROWS = [
    ("収入", "売上代金", "現金売上"),
    ("収入", "売上代金", "売掛金回収"),
    ("収入", "売上代金", "手形期日落"),
    ("収入", "売上代金", "手形割引"),
    ("収入", "前受金", "前受金"),
    ("財務", "借入金","借入金"),
    ("収入", "その他", "その他"),
    ("支出", "仕入代金", "現金仕入"),
    ("支出", "仕入代金", "買掛金支払"),
    ("支出", "仕入代金", "手形決済"),
    ("支出", "賃金及び給与", "賃金及び給与"),
    ("支出", "家賃", "家賃"),
    ("支出", "前渡金", "前渡金"),
    ("支出", "諸経費", "諸経費"),
    ("支出", "その他（設備等）", "その他（設備等）"),
    ("財務","借入金返済", "借入金返済"),
]

if "input_rows" not in st.session_state:
    st.session_state["input_rows"] = pd.DataFrame(
        columns=["月", "区分", "大分類", "小分類", "金額"]
    )

if "amount_digits_bridge" not in st.session_state:
    st.session_state["amount_digits_bridge"] = ""

if "last_amount_bridge" not in st.session_state:
    st.session_state["last_amount_bridge"] = 0

def fmt_yen(x):
    try:
        return f"{int(x):,}"
    except Exception:
        return x

def append_bridge_digits(s: str):
    st.session_state["amount_digits_bridge"] += s

def clear_bridge_digits():
    st.session_state["amount_digits_bridge"] = ""

def backspace_bridge_digits():
    st.session_state["amount_digits_bridge"] = st.session_state["amount_digits_bridge"][:-1]

st.subheader("前回データの読込")

uploaded_prev = st.file_uploader(
    "前回保存した資金繰り表をアップロードしてください",
    type=["xlsx"],
    key="prev_cashflow_file"
)

if uploaded_prev is not None:
    if st.button("この資金繰り表を復元"):
        restored_start_month, carry_in_restored, restored_df = restore_cashflow_template(uploaded_prev)
        st.session_state["input_rows"] = restored_df
        st.session_state["carry_in_man"] = int(carry_in_restored / 10000)
        st.session_state["start_month_restore"] = restored_start_month
        st.success("前回の資金繰り表を復元しました。")
        st.rerun()

st.subheader("1. 入力")

if "start_month_restore" not in st.session_state:
    st.session_state["start_month_restore"] = "1月"

start_month = st.selectbox(
    "開始月",
    MONTHS,
    index=MONTHS.index(st.session_state["start_month_restore"])
)
start_idx = MONTHS.index(start_month)
target_months = [MONTHS[(start_idx + i) % 12] for i in range(6)]

if "carry_in_man" not in st.session_state:
    st.session_state["carry_in_man"] = 300

carry_in_man = st.number_input(
    "前月繰越（万円）",
    min_value=0,
    step=10,
    key="carry_in_man"
)
carry_in = carry_in_man * 10000

input_mode = st.radio("入力方法", ["個別入力", "全月共通"], horizontal=True)

c1, c2, c3, c4 = st.columns([1, 1, 1.2, 1.4])

with c1:
    if input_mode == "個別入力":
        month_val = st.selectbox("月", target_months, index=0)
    else:
        st.markdown("**対象月**")
        st.write(" / ".join(target_months))
        month_val = None

with c2:
    kubun_val = st.selectbox("区分", ["収入", "支出", "財務"], index=0)

with c3:
    major_options = list(CATEGORY_STRUCTURE[kubun_val].keys())
    major_val = st.selectbox("大分類", major_options)

with c4:
    minor_options = CATEGORY_STRUCTURE[kubun_val][major_val]
    minor_val = st.selectbox("小分類", minor_options)

st.markdown("### 金額入力")

unit_val = st.radio(
    "単位",
    ["万円", "千円", "円"],
    index=0,
    horizontal=True,
    key="bridge_unit"
)

unit_map = {"万円": 10000, "千円": 1000, "円": 1}

digits_now = st.session_state["amount_digits_bridge"] or "0"
st.text_input("入力中", value=digits_now, disabled=True)

key_rows = [
    ["7", "8", "9", "000"],
    ["4", "5", "6", "00"],
    ["1", "2", "3", "0"],
]

for r_idx, row in enumerate(key_rows):
    cols = st.columns(4)
    for i, key in enumerate(row):
        if cols[i].button(key, key=f"bridge_key_{r_idx}_{key}"):
            append_bridge_digits(key)

cols = st.columns(5)
if cols[0].button("C", key="bridge_clear"):
    clear_bridge_digits()
if cols[1].button("⌫", key="bridge_back"):
    backspace_bridge_digits()
if cols[2].button("+1", key="bridge_plus1"):
    st.session_state["amount_digits_bridge"] = str(int(st.session_state["amount_digits_bridge"] or "0") + 1)
if cols[3].button("+10", key="bridge_plus10"):
    st.session_state["amount_digits_bridge"] = str(int(st.session_state["amount_digits_bridge"] or "0") + 10)
if cols[4].button("前回金額を使う"):
    st.session_state["amount_digits_bridge"] = str(
        int(st.session_state["last_amount_bridge"] / unit_map[unit_val])
    )



amount_val = int(st.session_state["amount_digits_bridge"] or "0") * unit_map[unit_val]

st.markdown(f"**確定金額：{amount_val:,.0f} 円**")

if st.button("追加", type="primary"):
    if amount_val > 0:
        add_months = [month_val] if input_mode == "個別入力" else target_months

        new_rows = pd.DataFrame([
            {
                "月": m,
                "区分": kubun_val,
                "大分類": major_val,
                "小分類": minor_val,
                "金額": int(amount_val),
            }
            for m in add_months
        ])

        st.session_state["input_rows"] = pd.concat(
            [st.session_state["input_rows"], new_rows],
            ignore_index=True
        )

        st.session_state["last_amount_bridge"] = amount_val
        clear_bridge_digits()
        st.rerun()
    else:
        st.warning("金額を入力してください。")


st.subheader("2. 入力済み一覧")

if not st.session_state["input_rows"].empty:
    show_df = st.session_state["input_rows"].copy()
    show_df["金額"] = show_df["金額"].map(fmt_yen)
    st.dataframe(show_df, use_container_width=True)

    d1, d2 = st.columns([1, 4])

    with d1:
        delete_idx = st.number_input(
            "削除No.",
            min_value=0,
            max_value=len(st.session_state["input_rows"]) - 1,
            step=1,
            value=0
        )
        if st.button("指定行を削除"):
            st.session_state["input_rows"] = st.session_state["input_rows"].drop(
                index=int(delete_idx)
            ).reset_index(drop=True)
            st.rerun()

    with d2:
        if st.button("すべて削除"):
            st.session_state["input_rows"] = pd.DataFrame(
                columns=["月", "区分", "大分類", "小分類", "金額"]
            )
            st.rerun()
else:
    st.info("まだ入力がありません。")

st.subheader("3. 判断表")

summary = pd.DataFrame(index=[r[2] for r in DISPLAY_ROWS], columns=target_months).fillna(0)

if not st.session_state["input_rows"].empty:
    work = st.session_state["input_rows"].copy()
    grouped = work.groupby(["月", "小分類"], as_index=False)["金額"].sum()

    for _, row in grouped.iterrows():
        if row["小分類"] in summary.index and row["月"] in summary.columns:
            summary.loc[row["小分類"], row["月"]] += row["金額"]

income_rows = [
    "現金売上", "売掛金回収", "手形期日落", "手形割引",
    "前受金",  "その他"
]
expense_rows = [
    "現金仕入", "買掛金支払", "手形決済",
    "賃金及び給与", "家賃", "前渡金", "諸経費", "その他（設備等）"
]

income_total = summary.loc[income_rows].sum()
expense_total = summary.loc[expense_rows].sum()
finance_total = summary.loc["借入金"] - summary.loc["借入金返済"]
balance_row = income_total - expense_total 


display_table = summary.copy()
display_table.loc["収入合計"] = income_total
display_table.loc["支出合計"] = expense_total
display_table.loc["差引過不足"] = balance_row
display_table.loc["借入金"] = summary.loc["借入金"]
display_table.loc["借入金返済"] = summary.loc["借入金返済"]

carry_series = []
ending_series = []

current_carry = carry_in
for m in target_months:
    carry_series.append(current_carry)
    month_balance = balance_row[m]
    month_finance_in = summary.loc["借入金", m]
    month_finance_out = summary.loc["借入金返済", m]
    month_ending = current_carry + month_balance + month_finance_in - month_finance_out
    ending_series.append(month_ending)
    current_carry = month_ending

display_table.loc["前月繰越"] = pd.Series(carry_series, index=target_months)
display_table.loc["翌月繰越"] = pd.Series(ending_series, index=target_months)

row_order = (
    ["前月繰越"]
    + income_rows
    + ["収入合計"]
    + expense_rows
    + ["支出合計", "差引過不足"]
    + finance_rows
    + ["翌月繰越"]
)

display_table = display_table.reindex(row_order)

st.dataframe(display_table.style.format(fmt_yen), use_container_width=True)

output = io.BytesIO()

with pd.ExcelWriter(output, engine="openpyxl") as writer:
    wb = writer.book
    ws = wb.create_sheet("資金繰り表")

    # ----------------------------
    # 基本設定
    # ----------------------------
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    def set_border(cell, left=thin, right=thin, top=thin, bottom=thin):
        cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    sub_fill = PatternFill("solid", fgColor="F3F3F3")
    total_fill = PatternFill("solid", fgColor="FFF2CC")

    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # ----------------------------
    # 月列定義
    # 各月: 予想 / 実績
    # ----------------------------
    # A列: 大区分
    # B列: 中区分
    # C列: 小区分
    # D列以降: 月別（予想 / 実績）
    start_col = 4  # D列
    month_col_map = {}

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=start_col + len(target_months) * 2 - 1
    )

    ws["A1"] = "資金繰り表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 月見出し
    for i, month_name in enumerate(target_months):
        col_pred = start_col + i * 2
        col_act = col_pred + 1
        month_col_map[month_name] = {"予想": col_pred, "実績": col_act}

        ws.merge_cells(start_row=2, start_column=col_pred, end_row=2, end_column=col_act)
        cell = ws.cell(row=2, column=col_pred)
        cell.value = month_name
        cell.font = bold_font
        cell.alignment = center
        cell.fill = header_fill

        ws.cell(row=3, column=col_pred, value="予想")
        ws.cell(row=3, column=col_act, value="実績")
        ws.cell(row=3, column=col_pred).alignment = center
        ws.cell(row=3, column=col_act).alignment = center
        ws.cell(row=3, column=col_pred).fill = sub_fill
        ws.cell(row=3, column=col_act).fill = sub_fill
        ws.cell(row=3, column=col_pred).font = bold_font
        ws.cell(row=3, column=col_act).font = bold_font

    # 左側見出し
    ws["A3"] = "区分"
    ws["B3"] = "分類"
    ws["C3"] = "科目"
    for c in ["A3", "B3", "C3"]:
        ws[c].font = bold_font
        ws[c].alignment = center
        ws[c].fill = sub_fill

    # ----------------------------
    # 行構造
    # ----------------------------
    row_defs = [
        ("収入", "", "前月繰越"),
        ("収入", "売上代金", "現金売上"),
        ("収入", "売上代金", "売掛金回収"),
        ("収入", "売上代金", "手形期日落"),
        ("収入", "売上代金", "手形割引"),
        ("収入", "", "前受金"),
        ("収入", "", "その他"),
        ("収入", "", "収入合計(A)"),

        ("支出", "仕入代金", "現金仕入"),
        ("支出", "仕入代金", "買掛金支払"),
        ("支出", "仕入代金", "手形決済"),
        ("支出", "", "賃金及び給与"),
        ("支出", "", "家賃"),
        ("支出", "", "前渡金"),
        ("支出", "", "諸経費"),
        ("支出", "", "その他（設備等）"),
        ("支出", "", "支出合計(B)"),

        ("差引", "", "差引過不足(A)-(B)=(C)"),

        ("財務", "", "借入金返済(D)"),
        ("財務", "", "借入金(E)"),

        ("最終", "", "翌月繰越(C)-(D)+(E)"),
    ]

    start_row = 4
    row_map = {}

    for i, (kubun, bunrui, kamoku) in enumerate(row_defs, start=start_row):
        ws.cell(row=i, column=1, value=kubun)
        ws.cell(row=i, column=2, value=bunrui)
        ws.cell(row=i, column=3, value=kamoku)
        row_map[kamoku] = i

    # ----------------------------
    # 左側セル結合
    # ----------------------------
    ws.merge_cells(start_row=row_map["前月繰越"], end_row=row_map["収入合計(A)"], start_column=1, end_column=1)
    ws.merge_cells(start_row=row_map["現金売上"], end_row=row_map["手形割引"], start_column=2, end_column=2)

    ws.merge_cells(start_row=row_map["現金仕入"], end_row=row_map["支出合計(B)"], start_column=1, end_column=1)
    ws.merge_cells(start_row=row_map["現金仕入"], end_row=row_map["手形決済"], start_column=2, end_column=2)

    ws.merge_cells(start_row=row_map["差引過不足(A)-(B)=(C)"], end_row=row_map["差引過不足(A)-(B)=(C)"], start_column=1, end_column=2)
    ws.merge_cells(start_row=row_map["借入金返済(D)"], end_row=row_map["借入金(E)"], start_column=1, end_column=1)
    ws.merge_cells(start_row=row_map["借入金返済(D)"], end_row=row_map["借入金(E)"], start_column=2, end_column=2)
    ws.merge_cells(start_row=row_map["翌月繰越(C)-(D)+(E)"], end_row=row_map["翌月繰越(C)-(D)+(E)"], start_column=1, end_column=2)

    # 左側中央寄せ
    for row in range(2, row_map["翌月繰越(C)-(D)+(E)"] + 1):
        for col in range(1, 4):
            ws.cell(row=row, column=col).alignment = center

    # ----------------------------
    # 値入力（予想列のみ）
    # 実績列は空欄
    # ----------------------------
    kamoku_to_input = {
        "前受金": "前受金",
        "その他": "その他収入",
        "現金売上": "現金売上",
        "売掛金回収": "売掛金回収",
        "手形期日落": "手形期日落",
        "手形割引": "手形割引",
        "現金仕入": "現金仕入",
        "買掛金支払": "買掛金支払",
        "手形決済": "手形決済",
        "賃金及び給与": "賃金及び給与",
        "家賃": "家賃",
        "前渡金": "前渡金",
        "諸経費": "諸経費",
        "その他（設備等）": "その他（設備等）",
        "借入金返済(D)": "借入金返済",
        "借入金(E)": "借入金",
    }

    # 初月前月繰越
    first_pred_col = month_col_map[target_months[0]]["予想"]
    ws.cell(row=row_map["前月繰越"], column=first_pred_col, value=carry_in)

    # 2か月目以降 前月繰越 = 前月翌月繰越
    for idx in range(1, len(target_months)):
        prev_month = target_months[idx - 1]
        this_month = target_months[idx]
        prev_end_col = month_col_map[prev_month]["予想"]
        this_carry_col = month_col_map[this_month]["予想"]
        ws.cell(
            row=row_map["前月繰越"],
            column=this_carry_col,
            value=f"={get_column_letter(prev_end_col)}{row_map['翌月繰越(C)-(D)+(E)']}"
        )

    # 入力済み明細を月別集計して予想列へ
    work = st.session_state["input_rows"].copy()
    if not work.empty:
        grouped = work.groupby(["月", "小分類"], as_index=False)["金額"].sum()
        for _, r in grouped.iterrows():
            m = r["月"]
            item = r["小分類"]
            if m in month_col_map:
                pred_col = month_col_map[m]["予想"]

                for display_name, source_name in kamoku_to_input.items():
                    if item == source_name:
                        ws.cell(row=row_map[display_name], column=pred_col, value=float(r["金額"]))

    # ----------------------------
    # 数式設定
    # ----------------------------
    income_formula_rows = ["現金売上", "売掛金回収", "手形期日落", "手形割引", "前受金", "その他"]
    expense_formula_rows = ["現金仕入", "買掛金支払", "手形決済", "賃金及び給与", "家賃", "前渡金", "諸経費", "その他（設備等）"]

    for m in target_months:
        pred_col = month_col_map[m]["予想"]
        col_letter = get_column_letter(pred_col)

        # 収入合計(A)
        income_formula = "+".join([f"{col_letter}{row_map[r]}" for r in income_formula_rows])
        ws.cell(row=row_map["収入合計(A)"], column=pred_col, value=f"={income_formula}")

        # 支出合計(B)
        expense_formula = "+".join([f"{col_letter}{row_map[r]}" for r in expense_formula_rows])
        ws.cell(row=row_map["支出合計(B)"], column=pred_col, value=f"={expense_formula}")

        # 差引過不足(A)-(B)=(C)
        ws.cell(
            row=row_map["差引過不足(A)-(B)=(C)"],
            column=pred_col,
            value=f"={col_letter}{row_map['収入合計(A)']}-{col_letter}{row_map['支出合計(B)']}"
        )

        # 翌月繰越(C)-(D)+(E)
        ws.cell(
            row=row_map["翌月繰越(C)-(D)+(E)"],
            column=pred_col,
            value=(
                f"={col_letter}{row_map['前月繰越']}"
                f"+{col_letter}{row_map['差引過不足(A)-(B)=(C)']}"
                f"-{col_letter}{row_map['借入金返済(D)']}"
                f"+{col_letter}{row_map['借入金(E)']}"
            )
        )

    # ----------------------------
    # 書式
    # ----------------------------
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 24

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 32

    for col in range(start_col, start_col + len(target_months) * 2):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # 金額書式
    for row in range(4, row_map["翌月繰越(C)-(D)+(E)"] + 1):
        for col in range(start_col, start_col + len(target_months) * 2):
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).alignment = right

    # 合計行色
    for name in ["収入合計(A)", "支出合計(B)", "差引過不足(A)-(B)=(C)", "翌月繰越(C)-(D)+(E)"]:
        r = row_map[name]
        for c in range(1, start_col + len(target_months) * 2):
            ws.cell(row=r, column=c).fill = total_fill
            ws.cell(row=r, column=c).font = bold_font

    # 罫線
    last_col = start_col + len(target_months) * 2 - 1
    last_row = row_map["翌月繰越(C)-(D)+(E)"]

    for r in range(2, last_row + 1):
        for c in range(1, last_col + 1):
            set_border(ws.cell(r, c))

    r = row_map["前月繰越"]

    for c in range(1, last_col + 1):
        ws.cell(r, c).border = Border(
            left=ws.cell(r, c).border.left,
            right=ws.cell(r, c).border.right,
            top=ws.cell(r, c).border.top,
            bottom=medium
        )
    
    # 行高さ（約2.5倍）
    for r in range(1, last_row + 1):
        ws.row_dimensions[r].height = 30

    # 外枠強調
    for c in range(1, last_col + 1):
        ws.cell(2, c).border = Border(
            left=ws.cell(2, c).border.left,
            right=ws.cell(2, c).border.right,
            top=medium,
            bottom=ws.cell(2, c).border.bottom
        )
        ws.cell(last_row, c).border = Border(
            left=ws.cell(last_row, c).border.left,
            right=ws.cell(last_row, c).border.right,
            top=ws.cell(last_row, c).border.top,
            bottom=medium
        )

    for r in range(2, last_row + 1):
        ws.cell(r, 1).border = Border(
            left=medium,
            right=ws.cell(r, 1).border.right,
            top=ws.cell(r, 1).border.top,
            bottom=ws.cell(r, 1).border.bottom
        )
        ws.cell(r, last_col).border = Border(
            left=ws.cell(r, last_col).border.left,
            right=medium,
            top=ws.cell(r, last_col).border.top,
            bottom=ws.cell(r, last_col).border.bottom
        )

    # 印刷設定
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:3"

excel_data = export_cashflow_template(
    input_rows=st.session_state["input_rows"],
    carry_in=carry_in,
    start_month=start_month
)

st.download_button(
    "資金繰り表をExcelでダウンロード",
    data=excel_data.getvalue(),
    file_name=f"資金繰り表_{datetime.now().strftime('%Y%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


st.subheader("4. 残高グラフ")

fig, ax = plt.subplots(figsize=(10, 4))

ax.plot(target_months, ending_series, marker="o")

from matplotlib.ticker import FuncFormatter

def yen_formatter(x, pos):
    if abs(x) >= 100_000_000:
        return f"{x/100_000_000:.1f}億円"
    elif abs(x) >= 10_000:
        return f"{int(x/10_000)}万円"
    else:
        return f"{int(x)}円"

ax.yaxis.set_major_formatter(FuncFormatter(yen_formatter))

ax.axhline(y=0, linestyle="--")
ax.set_title("翌月繰越残高の推移")
ax.set_ylabel("")
ax.grid(True, alpha=0.3)

st.pyplot(fig)
